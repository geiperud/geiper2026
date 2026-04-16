"""
forms_to_json.py
Convierte las respuestas aprobadas del Microsoft Forms de GEIPER
(Excel en SharePoint) al formato eventos_feed.json del sitio web.

USO:
1. Descarga el Excel desde SharePoint y guárdalo en la misma carpeta que este script
2. Abre el archivo y agrega una columna "Aprobado" con "SI" en las filas que quieras publicar
3. Ejecuta: python forms_to_json.py
4. Se genera/actualiza el archivo eventos_feed.json
5. Haz commit y push al repo de GEIPER

REQUISITOS:
    pip install openpyxl
"""

import json
import os
from datetime import datetime
import openpyxl

# ── Configuración ──────────────────────────────────────────────
EXCEL_FILE = "Registro de Eventos GEIPER.xlsx"
OUTPUT_JSON = "eventos_feed.json"

# Mapeo: nombre columna en Excel → clave en JSON
COLUMN_MAP = {
    "Título del evento":                                    "title",
    "Fecha del evento":                                     "date",
    "Tipo de evento":                                       "type",
    "Modalidad":                                            "modalidad",
    "Descripción (máx. 300 caracteres)":                   "description",
    "Link del evento (Pagina web o publicaciones relacionadas)": "link",
    "Aprobado":                                             "aprobado",  # columna manual que tú agregas
}

# Tipos de evento normalizados
TYPE_NORMALIZE = {
    "congreso":    "Congreso",
    "seminario":   "Seminario",
    "taller":      "Taller",
    "webinar":     "Webinar",
    "conferencia": "Conferencia",
    "otro":        "Evento",
}

def normalize_type(raw_type: str) -> str:
    if not raw_type:
        return "Evento"
    return TYPE_NORMALIZE.get(raw_type.strip().lower(), raw_type.strip().capitalize())

def normalize_date(raw_date) -> str:
    """Convierte fecha de Excel a formato YYYY-MM-DD"""
    if isinstance(raw_date, datetime):
        return raw_date.strftime("%Y-%m-%d")
    if isinstance(raw_date, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(raw_date.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return str(raw_date)

def build_tags(event_type: str, modalidad: str) -> list:
    """Genera tags automáticamente desde tipo y modalidad"""
    tags = []
    if event_type:
        tags.append(normalize_type(event_type))
    if modalidad:
        tags.append(modalidad.strip().capitalize())
    return tags

def load_existing_json(path: str) -> list:
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ No se encontró '{EXCEL_FILE}' en esta carpeta.")
        print("   Descárgalo desde SharePoint y colócalo aquí.")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    # Leer encabezados
    headers = [cell.value for cell in ws[1]]
    print(f"📋 Columnas encontradas: {headers}\n")

    # Mapear posiciones de columnas
    col_index = {}
    for col_name, json_key in COLUMN_MAP.items():
        try:
            col_index[json_key] = headers.index(col_name)
        except ValueError:
            if json_key != "aprobado":
                print(f"⚠️  Columna no encontrada: '{col_name}' — se omitirá.")

    if "aprobado" not in col_index:
        print("⚠️  No se encontró columna 'Aprobado'.")
        print("   Agrega manualmente una columna llamada 'Aprobado' en el Excel")
        print("   y escribe 'SI' en las filas que quieras publicar.\n")

    # Procesar filas
    nuevos_eventos = []
    filas_procesadas = 0
    filas_aprobadas = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        filas_procesadas += 1

        # Verificar aprobación
        if "aprobado" in col_index:
            aprobado = str(row[col_index["aprobado"]] or "").strip().upper()
            if aprobado != "SI":
                continue

        filas_aprobadas += 1

        # Extraer campos
        title       = str(row[col_index.get("title", 0)] or "").strip()
        raw_date    = row[col_index.get("date", 1)]
        event_type  = str(row[col_index.get("type", 2)] or "").strip()
        modalidad   = str(row[col_index.get("modalidad", 3)] or "").strip()
        description = str(row[col_index.get("description", 4)] or "").strip()
        link        = str(row[col_index.get("link", 5)] or "").strip()

        if not title:
            print(f"⚠️  Fila sin título, omitida.")
            continue

        evento = {
            "title":       title,
            "date":        normalize_date(raw_date),
            "type":        normalize_type(event_type),
            "modalidad":   modalidad.capitalize() if modalidad else "Por confirmar",
            "description": description[:300],  # respetar límite
            "link":        link if link.startswith("http") else "#",
            "tags":        build_tags(event_type, modalidad),
        }
        nuevos_eventos.append(evento)
        print(f"✅ Aprobado: {title} ({evento['date']})")

    print(f"\n📊 Resumen: {filas_procesadas} respuestas leídas, {filas_aprobadas} aprobadas.")

    if not nuevos_eventos:
        print("ℹ️  No hay eventos nuevos para agregar.")
        return

    # Ordenar por fecha (más próximos primero)
    nuevos_eventos.sort(key=lambda e: e["date"])

    # Guardar JSON
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(nuevos_eventos, f, ensure_ascii=False, indent=2)

    print(f"\n🎉 {OUTPUT_JSON} actualizado con {len(nuevos_eventos)} eventos.")
    print("   Ahora haz: git add . && git commit -m 'actualizar eventos' && git push")

if __name__ == "__main__":
    main()
